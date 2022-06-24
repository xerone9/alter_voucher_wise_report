import openpyxl as xl
import os
import glob
import csv
from xlsxwriter.workbook import Workbook


def listToStringWithoutBrackets(value):
    return str(value).replace('[','').replace(']','').replace('(','').replace(')','').replace(',','').replace("'",'').replace(' ','')


numbers = {
        "Jan": "1-",
        "Feb": "2-",
        "Mar": "3-",
        "Apr": "4-",
        "May": "5-",
        "Jun": "6-",
        "Jul": "7-",
        "Aug": "8-",
        "Sep": "9-",
        "Oct": "10-",
        "Nov": "11-",
        "Dec": "12-",
        }
months = [
        "Jan",
        "Feb",
        "Mar",
        "Apr",
        "May",
        "Jun",
        "Jul",
        "Aug",
        "Sep",
        "Oct",
        "Nov",
        "Dec"
]

#################### Converting CSV file To xlsx ##################################

desktop = os.path.expanduser("~\Desktop\\")
for csvfile in glob.glob(os.path.join(desktop, 'voucher_wise_deposit.csv')):
    workbook = Workbook(csvfile[:-4] + '.xlsx')
    worksheet = workbook.add_worksheet()
    with open(csvfile, 'rt', encoding='utf8') as f:
        reader = csv.reader(f)
        for r, row in enumerate(reader):
            for c, col in enumerate(row):
                worksheet.write(r, c, col)
    workbook.close()


############################ Setting up xlsx File #######################################

desktop = os.path.expanduser("~\Desktop\\voucher_wise_deposit.xlsx")

wb = xl.load_workbook(desktop)
sheet = wb['Sheet1']
sheet.title = 'receipts_Grand'
sheet = wb['receipts_Grand']
cell = sheet.cell(1, 1)

cell = sheet.cell(1, 1)
cell.value = 'Date'
sheet.delete_cols(2)
cell = sheet.cell(1, 2)
cell.value = 'ID'
cell = sheet.cell(1, 3)
cell.value = 'Name'
cell = sheet.cell(1, 4)
cell.value = 'Narration'
cell = sheet.cell(1, 5)
cell.value = 'Receipts'
cell = sheet.cell(1, 6)
cell.value = 'Amount'
cell = sheet.cell(1, 7)
cell.value = 'Amount'

###################### Placing Receipt No Values in Column and fixing CSV File Date Issue########################

for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 4)
    txt = cell.value
    receipts = [int(s) for s in txt.split() if s.isdigit()]
    work = sheet.cell(row, 5)
    work.value = listToStringWithoutBrackets(receipts)
    amount = sheet.cell(row, 6)
    amount_fix = sheet.cell(row, 7)
    amount_fix.value = listToStringWithoutBrackets(amount.value)
    stu_id = sheet.cell(row, 2)
    stu_id2 = str(stu_id.value)
    month = stu_id2[0] + stu_id2[1] + stu_id2[2]
    monthAlter = stu_id2[-3] + stu_id2[-2] + stu_id2[-1]
    words = stu_id2.split("-")

    if monthAlter in months:
        stu_id2 = str(monthAlter + stu_id.value[2] + stu_id.value[0] + stu_id.value[1])
        words = stu_id2.split("-")
    output = ""
    for word in words:
        output += numbers.get(word, word)

    acid = str(output)
    wordeg = list(acid)

    wordeg.insert(-4, '-')
    acid1 = listToStringWithoutBrackets(wordeg)
    stu_id.value = stu_id2
    if monthAlter in months:
        stu_id2 = str(monthAlter + stu_id.value[2] + stu_id.value[0] + stu_id.value[1])
        words = stu_id.value.split("-")
        output = ""
        for word in words:
            output += numbers.get(word, word)

        wordeg = list(output)
        wordeg.insert(-2, '20')
        acid1 = listToStringWithoutBrackets(wordeg)
        stu_id.value = acid1
    if month in months:
        stu_id2 = str(month + stu_id.value[-3] + stu_id.value[-2] + stu_id.value[-1])
        words = stu_id.value.split("-")
        output = ""
        for word in words:
            output += numbers.get(word, word)

        wordeg = list(output)
        wordeg.insert(-2, '20')
        acid1 = listToStringWithoutBrackets(wordeg)
        stu_id.value = acid1


sheet.delete_cols(6)
wb.save(desktop)
os.startfile(desktop)
